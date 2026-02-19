"""
SET Financial Corporation — Monthly EOM Close Automation
Main orchestrator: parses all source files, generates journal entries,
runs reconciliation, and produces the close package.

Usage:
    python run_close.py --month 2026-01 --data-dir ./data/sample/
"""
import sys
import os
from pathlib import Path
from decimal import Decimal
from datetime import date
import json

# Add project root to path
PROJECT_ROOT = Path(__file__).parent.parent
sys.path.insert(0, str(PROJECT_ROOT / "src"))

from parsers.collection_parser import parse_collection_register
from parsers.loan_register_parser import parse_loan_register
from parsers.charge_off_parser import parse_charge_offs
from parsers.unearned_parser import parse_unearned_register
from parsers.pier_statement_parser import parse_pier_statement
from parsers.dpv_statement_parser import parse_dpv_statement
from parsers.metacorp_parser import JANUARY_2026_SALE

from journal_entries.je_engine import (
    generate_je1_finance_income, generate_je2_insurance_earnings,
    generate_je3_originations, generate_je4_collections,
    generate_je5_charge_offs, generate_je6_bad_debt_sale,
    generate_je7_pier_interest, generate_je8_dpv_interest,
    generate_je9_allowance, generate_je10_recoveries,
    JournalEntry,
)

from reconciliation.recon_engine import (
    ReconciliationReport, build_loans_receivable_recon,
    build_unearned_interest_recon, build_pier_balance_recon,
    build_charge_off_validation, build_collections_validation,
)


# ──────────────────────────────────────────────────────────────
# QBO PRIOR PERIOD BALANCES (from Balance Sheet as of 12/31/2025)
# These would normally come from QBO API; hardcoded for Jan 2026
# Since the uploaded BS is as-of 2/16/2026 (YTD), we need to
# back into December 2025 ending balances.
# For now we use the current QBO values and will reconcile forward.
# ──────────────────────────────────────────────────────────────

# Current QBO balances (as of 2/16/2026 — effectively Jan 2026 close)
QBO_BALANCES = {
    "110001_gross": Decimal("35243907.50"),
    "110200_accum_co": Decimal("-33313354.46"),
    "110002_allowance": Decimal("-328194.02"),
    "110010_unearned_interest": Decimal("-55038.50"),
    "110050_unearned_life": Decimal("-1.28"),
    "110060_unearned_ah": Decimal("-0.71"),
    "110070_unearned_iui": Decimal("0"),
    "110080_unearned_prop": Decimal("0"),
    "110090_unearned_auto": Decimal("0"),
    "total_unearned_insurance": Decimal("-1.99"),
    "291300_pier_loc": Decimal("1591414.81"),
    "290060_accrued_expenses": Decimal("95349.58"),
}


def run_january_2026_close():
    """Execute the full January 2026 EOM close process."""
    data_dir = PROJECT_ROOT / "data" / "sample"
    output_dir = PROJECT_ROOT / "output" / "2026-01"
    output_dir.mkdir(parents=True, exist_ok=True)

    je_date = date(2026, 1, 31)
    period = "January 2026"

    print("=" * 70)
    print(f"  SET Financial Corporation — EOM Close: {period}")
    print("=" * 70)

    # ──────────────────────────────────────────────────────────
    # PHASE 1: PARSE ALL SOURCE FILES
    # ──────────────────────────────────────────────────────────
    print("\n📁 PHASE 1: Parsing Source Files...")

    # 1. Collection Register
    coll_path = data_dir / "Collection Register_January2026.xlsx"
    coll_df, coll_summary = parse_collection_register(str(coll_path))
    print(f"  ✓ Collection Register: {coll_summary.transaction_count} transactions, "
          f"${coll_summary.total_collected} collected")

    # 2. Loan Register
    loan_path = data_dir / "Loan Register_January2026.xlsx"
    loan_df, loan_summary = parse_loan_register(str(loan_path))
    print(f"  ✓ Loan Register: {loan_summary.loan_count} loans, "
          f"${loan_summary.note_amount} originated")

    # 3. Charge Offs
    co_path = data_dir / "Charge Offs_January2026.xlsx"
    co_df, co_summary = parse_charge_offs(str(co_path))
    print(f"  ✓ Charge Offs: {co_summary.account_count} accounts, "
          f"${co_summary.total_charge_off_amt} total")

    # 4. Unearned Register
    ue_path = data_dir / "Unearned_January2026.xlsx"
    ue_df, ue_summary = parse_unearned_register(str(ue_path))
    print(f"  ✓ Unearned Register: {ue_summary.loan_count} loans, "
          f"${ue_summary.total_unearned_interest} total unearned interest")

    # 5. Pier Statement
    pier_path = data_dir / "PAT LLC Series 47 - SET - January 2026.pdf"
    pier_data = parse_pier_statement(str(pier_path))
    print(f"  ✓ Pier Statement: Balance ${pier_data.ending_balance}, "
          f"Interest ${pier_data.accrued_interest_due}")

    # 6. DPV Statement
    dpv_path = data_dir / "2026 - DPV LLC - JAN.pdf"
    dpv_data = parse_dpv_statement(str(dpv_path))
    print(f"  ✓ DPV Statement: Balance ${dpv_data.principal_balance}, "
          f"Interest ${dpv_data.total_due}")

    # 7. Metacorp Sale
    metacorp = JANUARY_2026_SALE
    print(f"  ✓ Metacorp Sale: {metacorp.num_accounts} accounts, "
          f"${metacorp.transfer_amount} proceeds")

    # ──────────────────────────────────────────────────────────
    # PHASE 2: GENERATE JOURNAL ENTRIES
    # ──────────────────────────────────────────────────────────
    print("\n📝 PHASE 2: Generating Journal Entries...")

    journal_entries = []

    # JE-1: Finance Income Recognition
    # The earned interest = change in unearned from prior month to current month.
    # For January 2026 (first automated month), we don't have December's register.
    #
    # Approach: Use the interest collected from the Collection Register as the
    # earned interest for the month. For PC loans, interest collected represents
    # what was earned from payments. The unearned register's InterestCollectedMonth
    # field is $0 (possibly not populated for all loan types).
    #
    # From Collection Register: InterestCollected = $159,298.55
    # This represents the interest portion of all payments received in January.
    # For PC loans, this is the portion that was previously in unearned.
    # For NLS loans, this is simple interest earned on the outstanding balance.
    #
    # We use this as the JE-1 earned interest amount, flagged for review.

    earned_interest_estimate = coll_summary.interest_collected
    je1_note = ("FIRST AUTOMATED MONTH: Using Collection Register InterestCollected "
                f"(${coll_summary.interest_collected}) as earned interest proxy. "
                "Future months will use month-over-month unearned register delta. "
                "This should be cross-checked with the accountant's current method.")

    je1 = generate_je1_finance_income(
        prior_month_unearned_interest=ue_summary.total_unearned_interest + earned_interest_estimate,
        current_month_unearned_interest=ue_summary.total_unearned_interest,
        je_date=je_date,
    )
    je1.notes = je1_note
    je1.review_required = True
    journal_entries.append(je1)
    print(f"  ✓ JE-1 Finance Income: ${je1.total_debits} (estimated from interest collected)")

    # JE-2: Insurance Premium Earnings
    # Similar issue — need prior month for delta. Insurance balances are tiny.
    current_insurance = {
        "CreditLife": ue_summary.unearned_credit_life,
        "Disability": ue_summary.unearned_disability,
        "IUI": ue_summary.unearned_iui,
        "Property": ue_summary.unearned_property,
        "VSI": ue_summary.unearned_vsi,
    }
    # Use QBO balances as "prior" for first month
    prior_insurance = {
        "CreditLife": abs(QBO_BALANCES["110050_unearned_life"]) + Decimal("0"),  # approximate
        "Disability": abs(QBO_BALANCES["110060_unearned_ah"]) + Decimal("0"),
        "IUI": abs(QBO_BALANCES["110070_unearned_iui"]),
        "Property": abs(QBO_BALANCES["110080_unearned_prop"]),
        "VSI": abs(QBO_BALANCES["110090_unearned_auto"]),
    }
    je2 = generate_je2_insurance_earnings(prior_insurance, current_insurance, je_date)
    journal_entries.append(je2)
    print(f"  ✓ JE-2 Insurance Earnings: ${je2.total_debits}")

    # JE-3: Loan Originations
    je3 = generate_je3_originations(
        note_amount=loan_summary.note_amount,
        finance_charge=loan_summary.finance_charge,
        credit_life_premium=loan_summary.credit_life_premium,
        ah_premium=loan_summary.ah_premium,
        cash_to_borrower=loan_summary.cash_to_borrower,
        balance_renewed=loan_summary.balance_renewed,
        je_date=je_date,
    )
    journal_entries.append(je3)
    print(f"  ✓ JE-3 Originations: ${je3.total_debits} (${loan_summary.loan_count} loans)")

    # JE-4: Collections
    je4 = generate_je4_collections(
        cash_received=coll_summary.cash_received,
        principal=coll_summary.principal,
        interest_collected=coll_summary.interest_collected,
        interest_rebate=coll_summary.interest_rebate,
        balance_renewed=coll_summary.balance_renewed,
        late_fees=coll_summary.late_fees,
        nsf_fees=coll_summary.nsf_fees,
        amount_to_refund=coll_summary.amount_to_refund,
        insurance_rebate=coll_summary.insurance_rebate,
        recovery=coll_summary.recovery,
        je_date=je_date,
    )
    journal_entries.append(je4)
    print(f"  ✓ JE-4 Collections: ${je4.total_debits}")

    # JE-5: Charge-Offs
    je5 = generate_je5_charge_offs(
        total_charge_off_amt=co_summary.total_charge_off_amt,
        charge_off_amount=co_summary.charge_off_amount,
        pc_interest_rebate=co_summary.pc_interest_rebate,
        je_date=je_date,
    )
    journal_entries.append(je5)
    print(f"  ✓ JE-5 Charge-Offs: ${co_summary.total_charge_off_amt} total "
          f"(${co_summary.charge_off_amount} net, ${co_summary.pc_interest_rebate} unearned)")

    # JE-6: Bad Debt Sale
    je6 = generate_je6_bad_debt_sale(
        transfer_amount=metacorp.transfer_amount,
        je_date=je_date,
        num_accounts=metacorp.num_accounts,
        total_balance=metacorp.total_current_balance,
    )
    journal_entries.append(je6)
    print(f"  ✓ JE-6 Bad Debt Sale: ${metacorp.transfer_amount}")

    # JE-7: Pier Interest
    je7 = generate_je7_pier_interest(
        accrued_interest=pier_data.accrued_interest_due,
        je_date=je_date,
        ending_balance=pier_data.ending_balance,
    )
    journal_entries.append(je7)
    print(f"  ✓ JE-7 Pier Interest: ${pier_data.accrued_interest_due}")

    # JE-8: DPV Interest
    je8 = generate_je8_dpv_interest(
        interest_amount=dpv_data.total_due,
        je_date=je_date,
        principal_balance=dpv_data.principal_balance,
    )
    journal_entries.append(je8)
    print(f"  ✓ JE-8 DPV Interest: ${dpv_data.total_due}")

    # JE-9: Allowance for Credit Losses
    # Calculate suggested allowance based on net receivable
    net_receivable = ue_summary.total_current_balance
    current_allowance = abs(QBO_BALANCES["110002_allowance"])
    # Target: ~18% of net receivable (based on historical loss rates)
    target_pct = Decimal("18")
    target_allowance = (net_receivable * target_pct / Decimal("100")).quantize(Decimal("0.01"))
    adjustment = target_allowance - current_allowance

    je9 = generate_je9_allowance(
        adjustment_amount=adjustment,
        je_date=je_date,
        portfolio_balance=net_receivable,
        target_allowance_pct=target_pct,
    )
    journal_entries.append(je9)
    print(f"  ✓ JE-9 Allowance: Suggested adjustment ${adjustment} "
          f"(current ${current_allowance}, target ${target_allowance})")

    # JE-10: Recoveries — now consolidated into JE-4 (Collections)
    # Recovery is tracked in Collection Register and booked as part of JE-4.
    # Keeping JE-10 as a $0 placeholder for documentation purposes.
    if coll_summary.recovery > 0:
        print(f"  ℹ JE-10 Recoveries: ${coll_summary.recovery} (included in JE-4 Collections)")

    # ──────────────────────────────────────────────────────────
    # PHASE 3: RECONCILIATION
    # ──────────────────────────────────────────────────────────
    print("\n🔍 PHASE 3: Reconciliation Checks...")

    recon = ReconciliationReport(period=period)

    # Pier balance check
    pier_recon = build_pier_balance_recon(
        statement_balance=pier_data.ending_balance,
        qbo_pier_loc=QBO_BALANCES["291300_pier_loc"],
    )
    recon.items.append(pier_recon)
    print(f"  {'✓' if pier_recon.status == 'PASS' else '✗'} Pier Balance: "
          f"Statement ${pier_data.ending_balance} vs QBO ${QBO_BALANCES['291300_pier_loc']} "
          f"→ {pier_recon.status}")

    # Unearned interest check
    # NOTE: QBO balances are as of 2/16/2026 (mid-Feb), not 1/31/2026.
    # The register shows January month-end unearned. The difference includes
    # February activity (new originations adding unearned, payments earning down).
    # For proper recon, we'd need QBO balance as of 1/31/2026.
    ue_recon = build_unearned_interest_recon(
        register_unearned=ue_summary.total_unearned_interest,
        qbo_unearned=QBO_BALANCES["110010_unearned_interest"],
    )
    ue_recon.notes = (f"QBO balance is as of 2/16/2026, not 1/31/2026. "
                      f"Diff of ${abs(ue_recon.difference)} likely includes Feb activity. "
                      f"Re-run with 1/31 QBO trial balance for true recon.")
    recon.items.append(ue_recon)
    print(f"  ⚠ Unearned Interest: Register ${ue_summary.total_unearned_interest} "
          f"vs QBO(2/16) ${abs(QBO_BALANCES['110010_unearned_interest'])} "
          f"→ EXPECTED DIFF due to date mismatch (diff: ${ue_recon.difference})")

    # Charge-off validation
    co_valid = build_charge_off_validation(
        net_chargeoff=co_summary.charge_off_amount,
        unearned_reversed=co_summary.pc_interest_rebate,
        total_chargeoff=co_summary.total_charge_off_amt,
    )
    recon.add_validation(
        "Charge-Off Components (Net + Unearned = Total)",
        co_valid["status"] == "PASS",
        f"Net ${co_summary.charge_off_amount} + Unearned ${co_summary.pc_interest_rebate} "
        f"= ${co_valid['expected_total']} vs Total ${co_summary.total_charge_off_amt}"
    )
    print(f"  {'✓' if co_valid['status'] == 'PASS' else '✗'} Charge-Off Components: {co_valid['status']}")

    # JE balance checks
    all_balanced = True
    for je in journal_entries:
        if not je.is_balanced and len(je.lines) > 0:
            all_balanced = False
            recon.add_validation(
                f"{je.je_number} Balance Check",
                False,
                f"Debits ${je.total_debits} ≠ Credits ${je.total_credits}"
            )
            print(f"  ✗ {je.je_number}: UNBALANCED (DR ${je.total_debits} vs CR ${je.total_credits})")

    if all_balanced:
        recon.add_validation("All JEs Balanced", True, "All journal entries have matching debits and credits")
        print(f"  ✓ All journal entries balanced")

    # Duplicate loan check
    dupes = loan_df[loan_df.duplicated(subset=["LoanNumber"], keep=False)]
    has_dupes = len(dupes) > 0
    recon.add_validation(
        "No Duplicate Loan Numbers",
        not has_dupes,
        f"{len(dupes)} duplicates found" if has_dupes else "All loan numbers unique"
    )
    print(f"  {'✗' if has_dupes else '✓'} Duplicate Loan Check: "
          f"{'FAIL' if has_dupes else 'PASS'}")

    # Portfolio ID validation
    from config.portfolios import VALID_PORTFOLIO_IDS
    invalid_ports = set()
    for df_check in [loan_df, co_df]:
        port_col = "PortfolioID" if "PortfolioID" in df_check.columns else "PortfolioId"
        if port_col in df_check.columns:
            for pid in df_check[port_col].unique():
                if int(pid) not in VALID_PORTFOLIO_IDS and int(pid) != 0:
                    invalid_ports.add(int(pid))
    recon.add_validation(
        "Valid Portfolio IDs",
        len(invalid_ports) == 0,
        f"Invalid IDs: {invalid_ports}" if invalid_ports else "All portfolio IDs valid"
    )
    print(f"  {'✗' if invalid_ports else '✓'} Portfolio ID Check: "
          f"{'FAIL - ' + str(invalid_ports) if invalid_ports else 'PASS'}")

    # ──────────────────────────────────────────────────────────
    # PHASE 4: GENERATE OUTPUT
    # ──────────────────────────────────────────────────────────
    print("\n📊 PHASE 4: Generating Output Files...")

    # Generate Excel close package
    generate_excel_output(journal_entries, recon, coll_summary, loan_summary,
                          co_summary, ue_summary, pier_data, dpv_data, metacorp,
                          output_dir, period)

    print(f"\n{'=' * 70}")
    print(f"  EOM Close Complete: {period}")
    print(f"  Journal Entries: {len(journal_entries)}")
    print(f"  Recon Checks: {recon.pass_count} PASS, {recon.warning_count} WARNING, {recon.fail_count} FAIL")
    print(f"  Validations: {len(recon.validation_checks)}")
    print(f"  Output: {output_dir}")
    print(f"{'=' * 70}")

    return journal_entries, recon


def generate_excel_output(journal_entries, recon, coll_summary, loan_summary,
                          co_summary, ue_summary, pier_data, dpv_data, metacorp,
                          output_dir, period):
    """Generate the Excel close package."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    money_format = '#,##0.00'
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    review_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # ── Sheet 1: Executive Summary ──
    ws = wb.active
    ws.title = "Executive Summary"
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25

    row = 1
    ws.merge_cells('A1:D1')
    ws['A1'] = f"SET Financial Corporation — EOM Close Package: {period}"
    ws['A1'].font = Font(bold=True, size=14)
    row = 3

    # Key Metrics
    ws[f'A{row}'] = "KEY METRICS"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws[f'C{row}'].fill = header_fill
    row += 1

    metrics = [
        ("Loans Originated", float(loan_summary.note_amount), f"{loan_summary.loan_count} loans"),
        ("Cash Disbursed", float(loan_summary.cash_to_borrower), "Net to borrowers"),
        ("Collections Received", float(coll_summary.cash_received), f"{coll_summary.transaction_count} payments"),
        ("Principal Collected", float(coll_summary.principal), ""),
        ("Interest Collected", float(coll_summary.interest_collected), "Via unearned"),
        ("Late/NSF Fees", float(coll_summary.late_fees + coll_summary.nsf_fees), ""),
        ("Gross Charge-Offs", float(co_summary.total_charge_off_amt), f"{co_summary.account_count} accounts"),
        ("Net Charge-Offs (P&L)", float(co_summary.charge_off_amount), "Bad debt expense"),
        ("Unearned Reversed on COs", float(co_summary.pc_interest_rebate), ""),
        ("Bad Debt Sale (Metacorp)", float(metacorp.transfer_amount), f"{metacorp.num_accounts} accounts"),
        ("Pier Interest", float(pier_data.accrued_interest_due), f"Balance: ${pier_data.ending_balance}"),
        ("DPV Interest", float(dpv_data.total_due), f"Balance: ${dpv_data.principal_balance}"),
        ("Recoveries", float(coll_summary.recovery), "Charged-off account collections"),
        ("Refunds", float(coll_summary.amount_to_refund), "Customer refunds"),
    ]

    for label, amount, note in metrics:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = amount
        ws[f'B{row}'].number_format = money_format
        ws[f'C{row}'] = note
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = thin_border
        row += 1

    row += 1
    # Portfolio snapshot
    ws[f'A{row}'] = "PORTFOLIO SNAPSHOT"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws[f'C{row}'].fill = header_fill
    row += 1

    portfolio_metrics = [
        ("Active Loans (Unearned Register)", ue_summary.loan_count, ""),
        ("Total Current Balance", float(ue_summary.total_current_balance), "Nortridge"),
        ("Total Unearned Interest", float(ue_summary.total_unearned_interest), "QBO 110010"),
        ("Total Unearned Insurance", float(ue_summary.total_unearned_insurance), "QBO 110050-110090"),
        ("Pier Facility Balance", float(pier_data.ending_balance), "QBO 291300"),
        ("DPV LOC Balance", float(dpv_data.principal_balance), "QBO 31100"),
    ]
    for label, value, note in portfolio_metrics:
        ws[f'A{row}'] = label
        if isinstance(value, (int,)):
            ws[f'B{row}'] = value
        else:
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = money_format
        ws[f'C{row}'] = note
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = thin_border
        row += 1

    # ── Sheet 2: Journal Entries ──
    ws_je = wb.create_sheet("Journal Entries")
    ws_je.column_dimensions['A'].width = 10
    ws_je.column_dimensions['B'].width = 12
    ws_je.column_dimensions['C'].width = 35
    ws_je.column_dimensions['D'].width = 15
    ws_je.column_dimensions['E'].width = 15
    ws_je.column_dimensions['F'].width = 45
    ws_je.column_dimensions['G'].width = 15

    row = 1
    ws_je.merge_cells('A1:F1')
    ws_je['A1'] = f"Journal Entries — {period}"
    ws_je['A1'].font = Font(bold=True, size=14)
    row = 3

    for je in journal_entries:
        # JE Header
        ws_je[f'A{row}'] = je.je_number
        ws_je[f'A{row}'].font = Font(bold=True, size=11)
        ws_je[f'B{row}'] = je.description
        ws_je[f'B{row}'].font = Font(bold=True)
        ws_je.merge_cells(f'B{row}:F{row}')
        if je.review_required:
            ws_je[f'G{row}'] = "⚠ REVIEW"
            ws_je[f'G{row}'].fill = review_fill
            ws_je[f'G{row}'].font = Font(bold=True)
        row += 1

        ws_je[f'A{row}'] = f"Date: {je.je_date}"
        ws_je[f'B{row}'] = f"Source: {je.source_file}"
        row += 1

        # Column headers
        for col_idx, header in enumerate(["", "Account", "Account Name", "Debit", "Credit", "Memo"], 1):
            cell = ws_je.cell(row=row, column=col_idx, value=header)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = thin_border
        row += 1

        # Lines
        for line in je.lines:
            ws_je[f'B{row}'] = line.account_code
            ws_je[f'C{row}'] = line.account_name
            if line.debit > 0:
                ws_je[f'D{row}'] = float(line.debit)
                ws_je[f'D{row}'].number_format = money_format
            if line.credit > 0:
                ws_je[f'E{row}'] = float(line.credit)
                ws_je[f'E{row}'].number_format = money_format
            ws_je[f'F{row}'] = line.memo
            for col in ['B', 'C', 'D', 'E', 'F']:
                ws_je[f'{col}{row}'].border = thin_border
            row += 1

        # Totals
        ws_je[f'C{row}'] = "TOTALS"
        ws_je[f'C{row}'].font = Font(bold=True)
        ws_je[f'D{row}'] = float(je.total_debits)
        ws_je[f'D{row}'].number_format = money_format
        ws_je[f'D{row}'].font = Font(bold=True)
        ws_je[f'E{row}'] = float(je.total_credits)
        ws_je[f'E{row}'].number_format = money_format
        ws_je[f'E{row}'].font = Font(bold=True)
        balanced = je.is_balanced
        ws_je[f'F{row}'] = "✓ Balanced" if balanced else f"✗ Diff: ${je.total_debits - je.total_credits}"
        ws_je[f'F{row}'].fill = pass_fill if balanced else fail_fill
        for col in ['C', 'D', 'E', 'F']:
            ws_je[f'{col}{row}'].border = thin_border
        row += 1

        # Notes
        if je.notes:
            ws_je[f'B{row}'] = f"Notes: {je.notes}"
            ws_je[f'B{row}'].font = Font(italic=True, size=9)
            ws_je.merge_cells(f'B{row}:F{row}')
            row += 1

        row += 1  # Blank row between JEs

    # ── Sheet 3: Reconciliation ──
    ws_recon = wb.create_sheet("Reconciliation")
    ws_recon.column_dimensions['A'].width = 40
    ws_recon.column_dimensions['B'].width = 18
    ws_recon.column_dimensions['C'].width = 18
    ws_recon.column_dimensions['D'].width = 15
    ws_recon.column_dimensions['E'].width = 12
    ws_recon.column_dimensions['F'].width = 30

    row = 1
    ws_recon.merge_cells('A1:F1')
    ws_recon['A1'] = f"Reconciliation Report — {period}"
    ws_recon['A1'].font = Font(bold=True, size=14)
    row = 3

    # Balance Reconciliations
    ws_recon[f'A{row}'] = "BALANCE RECONCILIATIONS"
    ws_recon[f'A{row}'].font = header_font
    ws_recon[f'A{row}'].fill = header_fill
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws_recon[f'{col}{row}'].fill = header_fill
    row += 1

    headers = ["Check", "Source Value", "QBO Value", "Difference", "Status", "Notes"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_recon.cell(row=row, column=col_idx, value=h)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = thin_border
    row += 1

    for item in recon.items:
        ws_recon[f'A{row}'] = item.name
        ws_recon[f'B{row}'] = float(item.source_value)
        ws_recon[f'B{row}'].number_format = money_format
        ws_recon[f'C{row}'] = float(item.target_value)
        ws_recon[f'C{row}'].number_format = money_format
        ws_recon[f'D{row}'] = float(item.difference)
        ws_recon[f'D{row}'].number_format = money_format
        ws_recon[f'E{row}'] = item.status
        status_fill = pass_fill if item.status == "PASS" else (warn_fill if item.status == "WARNING" else fail_fill)
        ws_recon[f'E{row}'].fill = status_fill
        ws_recon[f'F{row}'] = item.notes
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_recon[f'{col}{row}'].border = thin_border
        row += 1

    row += 2

    # Validation Checks
    ws_recon[f'A{row}'] = "VALIDATION CHECKS"
    ws_recon[f'A{row}'].font = header_font
    ws_recon[f'A{row}'].fill = header_fill
    for col in ['B', 'C']:
        ws_recon[f'{col}{row}'].fill = header_fill
    row += 1

    for col_idx, h in enumerate(["Check", "Status", "Detail"], 1):
        cell = ws_recon.cell(row=row, column=col_idx, value=h)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = thin_border
    row += 1

    for check in recon.validation_checks:
        ws_recon[f'A{row}'] = check["name"]
        ws_recon[f'B{row}'] = check["status"]
        status_fill = pass_fill if check["status"] == "PASS" else fail_fill
        ws_recon[f'B{row}'].fill = status_fill
        ws_recon[f'C{row}'] = check.get("detail", "")
        for col in ['A', 'B', 'C']:
            ws_recon[f'{col}{row}'].border = thin_border
        row += 1

    # ── Sheet 4: Source Data Summary ──
    ws_data = wb.create_sheet("Source Data Summary")
    ws_data.column_dimensions['A'].width = 35
    ws_data.column_dimensions['B'].width = 18
    ws_data.column_dimensions['C'].width = 18

    row = 1
    ws_data['A1'] = f"Source Data Summary — {period}"
    ws_data['A1'].font = Font(bold=True, size=14)
    row = 3

    # Collection Register summary
    ws_data[f'A{row}'] = "COLLECTION REGISTER"
    ws_data[f'A{row}'].font = header_font
    ws_data[f'A{row}'].fill = header_fill
    ws_data[f'B{row}'].fill = header_fill
    row += 1
    coll_items = [
        ("Transactions", coll_summary.transaction_count),
        ("Total Collected", float(coll_summary.total_collected)),
        ("Principal", float(coll_summary.principal)),
        ("Interest Collected", float(coll_summary.interest_collected)),
        ("Interest Rebate", float(coll_summary.interest_rebate)),
        ("Late Fees", float(coll_summary.late_fees)),
        ("NSF Fees", float(coll_summary.nsf_fees)),
        ("Insurance Rebate", float(coll_summary.insurance_rebate)),
        ("Balance Renewed", float(coll_summary.balance_renewed)),
        ("Recovery", float(coll_summary.recovery)),
        ("Amount to Refund", float(coll_summary.amount_to_refund)),
        ("Cash Received", float(coll_summary.cash_received)),
    ]
    for label, val in coll_items:
        ws_data[f'A{row}'] = label
        ws_data[f'B{row}'] = val
        if isinstance(val, float):
            ws_data[f'B{row}'].number_format = money_format
        ws_data[f'A{row}'].border = thin_border
        ws_data[f'B{row}'].border = thin_border
        row += 1

    row += 1

    # Loan Register summary
    ws_data[f'A{row}'] = "LOAN REGISTER"
    ws_data[f'A{row}'].font = header_font
    ws_data[f'A{row}'].fill = header_fill
    ws_data[f'B{row}'].fill = header_fill
    row += 1
    loan_items = [
        ("Loans Originated", loan_summary.loan_count),
        ("Note Amount", float(loan_summary.note_amount)),
        ("Finance Charge", float(loan_summary.finance_charge)),
        ("Cash to Borrower", float(loan_summary.cash_to_borrower)),
        ("Credit Life Premium", float(loan_summary.credit_life_premium)),
        ("A&H Premium", float(loan_summary.ah_premium)),
        ("Balance Renewed", float(loan_summary.balance_renewed)),
    ]
    for label, val in loan_items:
        ws_data[f'A{row}'] = label
        ws_data[f'B{row}'] = val
        if isinstance(val, float):
            ws_data[f'B{row}'].number_format = money_format
        ws_data[f'A{row}'].border = thin_border
        ws_data[f'B{row}'].border = thin_border
        row += 1

    row += 1

    # Charge Off summary
    ws_data[f'A{row}'] = "CHARGE OFFS"
    ws_data[f'A{row}'].font = header_font
    ws_data[f'A{row}'].fill = header_fill
    ws_data[f'B{row}'].fill = header_fill
    row += 1
    co_items = [
        ("Accounts", co_summary.account_count),
        ("Note Amount", float(co_summary.note_amount)),
        ("Net Charge Off (P&L)", float(co_summary.charge_off_amount)),
        ("Unearned Reversed", float(co_summary.pc_interest_rebate)),
        ("Total Charge Off (Gross)", float(co_summary.total_charge_off_amt)),
    ]
    for label, val in co_items:
        ws_data[f'A{row}'] = label
        ws_data[f'B{row}'] = val
        if isinstance(val, float):
            ws_data[f'B{row}'].number_format = money_format
        ws_data[f'A{row}'].border = thin_border
        ws_data[f'B{row}'].border = thin_border
        row += 1

    row += 1

    # Unearned Register summary
    ws_data[f'A{row}'] = "UNEARNED REGISTER"
    ws_data[f'A{row}'].font = header_font
    ws_data[f'A{row}'].fill = header_fill
    ws_data[f'B{row}'].fill = header_fill
    row += 1
    ue_items = [
        ("Active Loans", ue_summary.loan_count),
        ("Current Balance (Portfolio)", float(ue_summary.total_current_balance)),
        ("Unearned Interest (New)", float(ue_summary.unearned_new_interest)),
        ("Unearned Interest (Existing)", float(ue_summary.unearned_existing_interest)),
        ("Total Unearned Interest", float(ue_summary.total_unearned_interest)),
        ("Unearned Finance Fees", float(ue_summary.total_unearned_finance_fees)),
        ("Unearned Credit Life", float(ue_summary.unearned_credit_life)),
        ("Unearned Disability (A&H)", float(ue_summary.unearned_disability)),
        ("Unearned IUI", float(ue_summary.unearned_iui)),
        ("Unearned Property", float(ue_summary.unearned_property)),
        ("Unearned VSI (Auto)", float(ue_summary.unearned_vsi)),
        ("Total Unearned Insurance", float(ue_summary.total_unearned_insurance)),
        ("Interest Collected This Month", float(ue_summary.interest_collected_month)),
    ]
    for label, val in ue_items:
        ws_data[f'A{row}'] = label
        ws_data[f'B{row}'] = val
        if isinstance(val, float):
            ws_data[f'B{row}'].number_format = money_format
        ws_data[f'A{row}'].border = thin_border
        ws_data[f'B{row}'].border = thin_border
        row += 1

    # ── Sheet 5: QBO Import Format ──
    ws_qbo = wb.create_sheet("QBO Import")
    ws_qbo.column_dimensions['A'].width = 12
    ws_qbo.column_dimensions['B'].width = 12
    ws_qbo.column_dimensions['C'].width = 12
    ws_qbo.column_dimensions['D'].width = 35
    ws_qbo.column_dimensions['E'].width = 15
    ws_qbo.column_dimensions['F'].width = 15
    ws_qbo.column_dimensions['G'].width = 40

    row = 1
    ws_qbo['A1'] = "QBO Journal Entry Import Format"
    ws_qbo['A1'].font = Font(bold=True, size=14)
    row = 2
    ws_qbo['A2'] = "Copy these entries into QuickBooks Online manually or via CSV import"
    ws_qbo['A2'].font = Font(italic=True, size=10)
    row = 4

    qbo_headers = ["JE #", "Date", "Account #", "Account Name", "Debit", "Credit", "Memo"]
    for col_idx, h in enumerate(qbo_headers, 1):
        cell = ws_qbo.cell(row=row, column=col_idx, value=h)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = thin_border
    row += 1

    for je in journal_entries:
        for line in je.lines:
            ws_qbo[f'A{row}'] = je.je_number
            ws_qbo[f'B{row}'] = je.je_date.strftime("%m/%d/%Y")
            ws_qbo[f'C{row}'] = line.account_code
            ws_qbo[f'D{row}'] = line.account_name
            if line.debit > 0:
                ws_qbo[f'E{row}'] = float(line.debit)
                ws_qbo[f'E{row}'].number_format = money_format
            if line.credit > 0:
                ws_qbo[f'F{row}'] = float(line.credit)
                ws_qbo[f'F{row}'].number_format = money_format
            ws_qbo[f'G{row}'] = line.memo
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
                ws_qbo[f'{col}{row}'].border = thin_border
            row += 1
        row += 1  # Blank between JEs

    # Save
    output_path = output_dir / "SET_Financial_EOM_Close_January2026.xlsx"
    wb.save(str(output_path))
    print(f"  ✓ Close package saved: {output_path}")

    # Also generate a flat CSV for QBO import
    csv_path = output_dir / "qbo_journal_entries_january2026.csv"
    with open(csv_path, 'w') as f:
        f.write("JE_Number,Date,Account_Number,Account_Name,Debit,Credit,Memo\n")
        for je in journal_entries:
            for line in je.lines:
                debit_str = f"{line.debit}" if line.debit > 0 else ""
                credit_str = f"{line.credit}" if line.credit > 0 else ""
                memo = line.memo.replace(",", ";").replace('"', "'")
                f.write(f'{je.je_number},{je.je_date},{line.account_code},"{line.account_name}",{debit_str},{credit_str},"{memo}"\n')
    print(f"  ✓ QBO import CSV saved: {csv_path}")

    return output_path


if __name__ == "__main__":
    run_january_2026_close()
