"""
Excel close package generator for SET Financial Corporation EOM Close.
Extracted from run_close.py to keep the orchestrator focused on business logic.
"""
from pathlib import Path


def generate_excel_output(journal_entries, recon, coll_summary, loan_summary,
                          co_summary, ue_summary, pier_data, dpv_data, metacorp,
                          output_dir, period):
    """Generate the Excel close package."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # Styles
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    subheader_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
    money_format = '#,##0.00'
    pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    warn_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )
    review_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    # ── Sheet 1: Executive Summary ──
    ws = wb.active
    ws.title = "Executive Summary"
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25

    row = 1
    ws.merge_cells('A1:D1')
    ws['A1'] = f"SET Financial Corporation — EOM Close Package: {period}"
    ws['A1'].font = Font(bold=True, size=14)
    row = 3

    # Key Metrics
    ws[f'A{row}'] = "KEY METRICS"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws[f'C{row}'].fill = header_fill
    row += 1

    metrics = [
        ("Loans Originated", float(loan_summary.note_amount), f"{loan_summary.loan_count} loans"),
        ("Cash Disbursed", float(loan_summary.cash_to_borrower), "Net to borrowers"),
        ("Collections Received", float(coll_summary.cash_received), f"{coll_summary.transaction_count} payments"),
        ("Principal Collected", float(coll_summary.principal), ""),
        ("Interest Collected", float(coll_summary.interest_collected), "Via unearned"),
        ("Late/NSF Fees", float(coll_summary.late_fees + coll_summary.nsf_fees), ""),
        ("Gross Charge-Offs", float(co_summary.total_charge_off_amt), f"{co_summary.account_count} accounts"),
        ("Net Charge-Offs (P&L)", float(co_summary.charge_off_amount), "Bad debt expense"),
        ("Unearned Reversed on COs", float(co_summary.pc_interest_rebate), ""),
        ("Bad Debt Sale (Metacorp)", float(metacorp.transfer_amount), f"{metacorp.num_accounts} accounts"),
        ("Pier Interest", float(pier_data.accrued_interest_due), f"Balance: ${pier_data.ending_balance}"),
        ("DPV Interest", float(dpv_data.total_due), f"Balance: ${dpv_data.principal_balance}"),
        ("Recoveries", float(coll_summary.recovery), "Charged-off account collections"),
        ("Refunds", float(coll_summary.amount_to_refund), "Customer refunds"),
    ]

    for label, amount, note in metrics:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = amount
        ws[f'B{row}'].number_format = money_format
        ws[f'C{row}'] = note
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = thin_border
        row += 1

    row += 1
    # Portfolio snapshot
    ws[f'A{row}'] = "PORTFOLIO SNAPSHOT"
    ws[f'A{row}'].font = header_font
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws[f'C{row}'].fill = header_fill
    row += 1

    portfolio_metrics = [
        ("Active Loans (Unearned Register)", ue_summary.loan_count, ""),
        ("Total Current Balance", float(ue_summary.total_current_balance), "Nortridge"),
        ("Total Unearned Interest", float(ue_summary.total_unearned_interest), "QBO 110010"),
        ("Total Unearned Insurance", float(ue_summary.total_unearned_insurance), "QBO 110050-110090"),
        ("Pier Facility Balance", float(pier_data.ending_balance), "QBO 291300"),
        ("DPV LOC Balance", float(dpv_data.principal_balance), "QBO 31100"),
    ]
    for label, value, note in portfolio_metrics:
        ws[f'A{row}'] = label
        if isinstance(value, (int,)):
            ws[f'B{row}'] = value
        else:
            ws[f'B{row}'] = value
            ws[f'B{row}'].number_format = money_format
        ws[f'C{row}'] = note
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = thin_border
        row += 1

    # ── Sheet 2: Journal Entries ──
    ws_je = wb.create_sheet("Journal Entries")
    ws_je.column_dimensions['A'].width = 10
    ws_je.column_dimensions['B'].width = 12
    ws_je.column_dimensions['C'].width = 35
    ws_je.column_dimensions['D'].width = 10
    ws_je.column_dimensions['E'].width = 15
    ws_je.column_dimensions['F'].width = 15
    ws_je.column_dimensions['G'].width = 45
    ws_je.column_dimensions['H'].width = 15

    row = 1
    ws_je.merge_cells('A1:G1')
    ws_je['A1'] = f"Journal Entries — {period}"
    ws_je['A1'].font = Font(bold=True, size=14)
    row = 3

    for je in journal_entries:
        # JE Header
        ws_je[f'A{row}'] = je.je_number
        ws_je[f'A{row}'].font = Font(bold=True, size=11)
        ws_je[f'B{row}'] = je.description
        ws_je[f'B{row}'].font = Font(bold=True)
        ws_je.merge_cells(f'B{row}:G{row}')
        if je.review_required:
            ws_je[f'H{row}'] = "⚠ REVIEW"
            ws_je[f'H{row}'].fill = review_fill
            ws_je[f'H{row}'].font = Font(bold=True)
        row += 1

        ws_je[f'A{row}'] = f"Date: {je.je_date}"
        ws_je[f'B{row}'] = f"Source: {je.source_file}"
        row += 1

        # Column headers
        for col_idx, header in enumerate(["", "Account", "Account Name", "Class", "Debit", "Credit", "Memo"], 1):
            cell = ws_je.cell(row=row, column=col_idx, value=header)
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.border = thin_border
        row += 1

        # Lines
        for line in je.lines:
            ws_je[f'B{row}'] = line.account_code
            ws_je[f'C{row}'] = line.account_name
            ws_je[f'D{row}'] = line.class_name
            if line.debit > 0:
                ws_je[f'E{row}'] = float(line.debit)
                ws_je[f'E{row}'].number_format = money_format
            if line.credit > 0:
                ws_je[f'F{row}'] = float(line.credit)
                ws_je[f'F{row}'].number_format = money_format
            ws_je[f'G{row}'] = line.memo
            for col in ['B', 'C', 'D', 'E', 'F', 'G']:
                ws_je[f'{col}{row}'].border = thin_border
            row += 1

        # Totals
        ws_je[f'C{row}'] = "TOTALS"
        ws_je[f'C{row}'].font = Font(bold=True)
        ws_je[f'E{row}'] = float(je.total_debits)
        ws_je[f'E{row}'].number_format = money_format
        ws_je[f'E{row}'].font = Font(bold=True)
        ws_je[f'F{row}'] = float(je.total_credits)
        ws_je[f'F{row}'].number_format = money_format
        ws_je[f'F{row}'].font = Font(bold=True)
        balanced = je.is_balanced
        ws_je[f'G{row}'] = "✓ Balanced" if balanced else f"✗ Diff: ${je.total_debits - je.total_credits}"
        ws_je[f'G{row}'].fill = pass_fill if balanced else fail_fill
        for col in ['C', 'E', 'F', 'G']:
            ws_je[f'{col}{row}'].border = thin_border
        row += 1

        # Notes
        if je.notes:
            ws_je[f'B{row}'] = f"Notes: {je.notes}"
            ws_je[f'B{row}'].font = Font(italic=True, size=9)
            ws_je.merge_cells(f'B{row}:G{row}')
            row += 1

        row += 1  # Blank row between JEs

    # ── Sheet 3: Reconciliation ──
    ws_recon = wb.create_sheet("Reconciliation")
    ws_recon.column_dimensions['A'].width = 40
    ws_recon.column_dimensions['B'].width = 18
    ws_recon.column_dimensions['C'].width = 18
    ws_recon.column_dimensions['D'].width = 15
    ws_recon.column_dimensions['E'].width = 12
    ws_recon.column_dimensions['F'].width = 30

    row = 1
    ws_recon.merge_cells('A1:F1')
    ws_recon['A1'] = f"Reconciliation Report — {period}"
    ws_recon['A1'].font = Font(bold=True, size=14)
    row = 3

    # Balance Reconciliations
    ws_recon[f'A{row}'] = "BALANCE RECONCILIATIONS"
    ws_recon[f'A{row}'].font = header_font
    ws_recon[f'A{row}'].fill = header_fill
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws_recon[f'{col}{row}'].fill = header_fill
    row += 1

    headers = ["Check", "Source Value", "QBO Value", "Difference", "Status", "Notes"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws_recon.cell(row=row, column=col_idx, value=h)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = thin_border
    row += 1

    for item in recon.items:
        ws_recon[f'A{row}'] = item.name
        ws_recon[f'B{row}'] = float(item.source_value)
        ws_recon[f'B{row}'].number_format = money_format
        ws_recon[f'C{row}'] = float(item.target_value)
        ws_recon[f'C{row}'].number_format = money_format
        ws_recon[f'D{row}'] = float(item.difference)
        ws_recon[f'D{row}'].number_format = money_format
        ws_recon[f'E{row}'] = item.status
        status_fill = pass_fill if item.status == "PASS" else (warn_fill if item.status == "WARNING" else fail_fill)
        ws_recon[f'E{row}'].fill = status_fill
        ws_recon[f'F{row}'] = item.notes
        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws_recon[f'{col}{row}'].border = thin_border
        row += 1

    row += 2

    # Validation Checks
    ws_recon[f'A{row}'] = "VALIDATION CHECKS"
    ws_recon[f'A{row}'].font = header_font
    ws_recon[f'A{row}'].fill = header_fill
    for col in ['B', 'C']:
        ws_recon[f'{col}{row}'].fill = header_fill
    row += 1

    for col_idx, h in enumerate(["Check", "Status", "Detail"], 1):
        cell = ws_recon.cell(row=row, column=col_idx, value=h)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = thin_border
    row += 1

    for check in recon.validation_checks:
        ws_recon[f'A{row}'] = check["name"]
        ws_recon[f'B{row}'] = check["status"]
        status_fill = pass_fill if check["status"] == "PASS" else fail_fill
        ws_recon[f'B{row}'].fill = status_fill
        ws_recon[f'C{row}'] = check.get("detail", "")
        for col in ['A', 'B', 'C']:
            ws_recon[f'{col}{row}'].border = thin_border
        row += 1

    # ── Sheet 4: Source Data Summary ──
    ws_data = wb.create_sheet("Source Data Summary")
    ws_data.column_dimensions['A'].width = 35
    ws_data.column_dimensions['B'].width = 18
    ws_data.column_dimensions['C'].width = 18

    row = 1
    ws_data['A1'] = f"Source Data Summary — {period}"
    ws_data['A1'].font = Font(bold=True, size=14)
    row = 3

    # Collection Register summary
    ws_data[f'A{row}'] = "COLLECTION REGISTER"
    ws_data[f'A{row}'].font = header_font
    ws_data[f'A{row}'].fill = header_fill
    ws_data[f'B{row}'].fill = header_fill
    row += 1
    coll_items = [
        ("Transactions", coll_summary.transaction_count),
        ("Total Collected", float(coll_summary.total_collected)),
        ("Principal", float(coll_summary.principal)),
        ("Interest Collected", float(coll_summary.interest_collected)),
        ("Interest Rebate", float(coll_summary.interest_rebate)),
        ("Late Fees", float(coll_summary.late_fees)),
        ("NSF Fees", float(coll_summary.nsf_fees)),
        ("Insurance Rebate", float(coll_summary.insurance_rebate)),
        ("Balance Renewed", float(coll_summary.balance_renewed)),
        ("Recovery", float(coll_summary.recovery)),
        ("Amount to Refund", float(coll_summary.amount_to_refund)),
        ("Cash Received", float(coll_summary.cash_received)),
    ]
    for label, val in coll_items:
        ws_data[f'A{row}'] = label
        ws_data[f'B{row}'] = val
        if isinstance(val, float):
            ws_data[f'B{row}'].number_format = money_format
        ws_data[f'A{row}'].border = thin_border
        ws_data[f'B{row}'].border = thin_border
        row += 1

    row += 1

    # Loan Register summary
    ws_data[f'A{row}'] = "LOAN REGISTER"
    ws_data[f'A{row}'].font = header_font
    ws_data[f'A{row}'].fill = header_fill
    ws_data[f'B{row}'].fill = header_fill
    row += 1
    loan_items = [
        ("Loans Originated", loan_summary.loan_count),
        ("Note Amount", float(loan_summary.note_amount)),
        ("Finance Charge", float(loan_summary.finance_charge)),
        ("Cash to Borrower", float(loan_summary.cash_to_borrower)),
        ("Credit Life Premium", float(loan_summary.credit_life_premium)),
        ("A&H Premium", float(loan_summary.ah_premium)),
        ("Balance Renewed", float(loan_summary.balance_renewed)),
    ]
    for label, val in loan_items:
        ws_data[f'A{row}'] = label
        ws_data[f'B{row}'] = val
        if isinstance(val, float):
            ws_data[f'B{row}'].number_format = money_format
        ws_data[f'A{row}'].border = thin_border
        ws_data[f'B{row}'].border = thin_border
        row += 1

    row += 1

    # Charge Off summary
    ws_data[f'A{row}'] = "CHARGE OFFS"
    ws_data[f'A{row}'].font = header_font
    ws_data[f'A{row}'].fill = header_fill
    ws_data[f'B{row}'].fill = header_fill
    row += 1
    co_items = [
        ("Accounts", co_summary.account_count),
        ("Note Amount", float(co_summary.note_amount)),
        ("Net Charge Off (P&L)", float(co_summary.charge_off_amount)),
        ("Unearned Reversed", float(co_summary.pc_interest_rebate)),
        ("Total Charge Off (Gross)", float(co_summary.total_charge_off_amt)),
    ]
    for label, val in co_items:
        ws_data[f'A{row}'] = label
        ws_data[f'B{row}'] = val
        if isinstance(val, float):
            ws_data[f'B{row}'].number_format = money_format
        ws_data[f'A{row}'].border = thin_border
        ws_data[f'B{row}'].border = thin_border
        row += 1

    row += 1

    # Unearned Register summary
    ws_data[f'A{row}'] = "UNEARNED REGISTER"
    ws_data[f'A{row}'].font = header_font
    ws_data[f'A{row}'].fill = header_fill
    ws_data[f'B{row}'].fill = header_fill
    row += 1
    ue_items = [
        ("Active Loans", ue_summary.loan_count),
        ("Current Balance (Portfolio)", float(ue_summary.total_current_balance)),
        ("Unearned Interest (New)", float(ue_summary.unearned_new_interest)),
        ("Unearned Interest (Existing)", float(ue_summary.unearned_existing_interest)),
        ("Total Unearned Interest", float(ue_summary.total_unearned_interest)),
        ("Unearned Finance Fees", float(ue_summary.total_unearned_finance_fees)),
        ("Unearned Credit Life", float(ue_summary.unearned_credit_life)),
        ("Unearned Disability (A&H)", float(ue_summary.unearned_disability)),
        ("Unearned IUI", float(ue_summary.unearned_iui)),
        ("Unearned Property", float(ue_summary.unearned_property)),
        ("Unearned VSI (Auto)", float(ue_summary.unearned_vsi)),
        ("Total Unearned Insurance", float(ue_summary.total_unearned_insurance)),
        ("Interest Collected This Month", float(ue_summary.interest_collected_month)),
    ]
    for label, val in ue_items:
        ws_data[f'A{row}'] = label
        ws_data[f'B{row}'] = val
        if isinstance(val, float):
            ws_data[f'B{row}'].number_format = money_format
        ws_data[f'A{row}'].border = thin_border
        ws_data[f'B{row}'].border = thin_border
        row += 1

    # ── Sheet 5: QBO Import Format ──
    ws_qbo = wb.create_sheet("QBO Import")
    ws_qbo.column_dimensions['A'].width = 12
    ws_qbo.column_dimensions['B'].width = 12
    ws_qbo.column_dimensions['C'].width = 12
    ws_qbo.column_dimensions['D'].width = 35
    ws_qbo.column_dimensions['E'].width = 10
    ws_qbo.column_dimensions['F'].width = 15
    ws_qbo.column_dimensions['G'].width = 15
    ws_qbo.column_dimensions['H'].width = 40

    row = 1
    ws_qbo['A1'] = "QBO Journal Entry Import Format"
    ws_qbo['A1'].font = Font(bold=True, size=14)
    row = 2
    ws_qbo['A2'] = "Copy these entries into QuickBooks Online manually or via CSV import"
    ws_qbo['A2'].font = Font(italic=True, size=10)
    row = 4

    qbo_headers = ["JE #", "Date", "Account #", "Account Name", "Class", "Debit", "Credit", "Memo"]
    for col_idx, h in enumerate(qbo_headers, 1):
        cell = ws_qbo.cell(row=row, column=col_idx, value=h)
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.border = thin_border
    row += 1

    for je in journal_entries:
        for line in je.lines:
            ws_qbo[f'A{row}'] = je.je_number
            ws_qbo[f'B{row}'] = je.je_date.strftime("%m/%d/%Y")
            ws_qbo[f'C{row}'] = line.account_code
            ws_qbo[f'D{row}'] = line.account_name
            ws_qbo[f'E{row}'] = line.class_name
            if line.debit > 0:
                ws_qbo[f'F{row}'] = float(line.debit)
                ws_qbo[f'F{row}'].number_format = money_format
            if line.credit > 0:
                ws_qbo[f'G{row}'] = float(line.credit)
                ws_qbo[f'G{row}'].number_format = money_format
            ws_qbo[f'H{row}'] = line.memo
            for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws_qbo[f'{col}{row}'].border = thin_border
            row += 1
        row += 1  # Blank between JEs

    # Save
    # Build dynamic filename from period
    period_parts = period.split()
    if len(period_parts) == 2:
        month_name, year = period_parts
        output_filename = f"SET_Financial_EOM_Close_{month_name}{year}.xlsx"
    else:
        output_filename = f"SET_Financial_EOM_Close_{period.replace(' ', '_')}.xlsx"

    output_path = Path(output_dir) / output_filename
    wb.save(str(output_path))
    print(f"  ✓ Close package saved: {output_path}")

    return output_path
